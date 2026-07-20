import uuid
import io
from datetime import date
from decimal import Decimal
from typing import Annotated

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.database import get_db
from app.core.dependencies import require_permission
from app.core import storage
from app.core.exceptions import NotFoundError, ValidationError
from app.models.finance import FinanceTransaction
from app.services import finance_service

router = APIRouter(prefix="/finances", tags=["Finanzas"])

_ALLOWED_MIME = {"image/jpeg", "image/png", "image/webp", "application/pdf"}
_MIME_TO_EXT = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "application/pdf": "pdf"}
_MAX_BYTES = 10 * 1024 * 1024  # 10 MB


def _serialize(t) -> dict:
    return {
        "id": str(t.id),
        "type": t.type,
        "category": t.category,
        "description": t.description,
        "amount_cordobas": float(t.amount_cordobas),
        "original_amount": float(t.original_amount) if t.original_amount else None,
        "original_currency": t.original_currency,
        "exchange_rate_used": float(t.exchange_rate_used) if t.exchange_rate_used else None,
        "patient": {"id": str(t.patient.id), "full_name": t.patient.full_name} if t.patient else None,
        "procedure": {"id": str(t.procedure.id), "name": t.procedure.name} if t.procedure else None,
        "operational_cost_snapshot": float(t.operational_cost_snapshot) if t.operational_cost_snapshot else None,
        "invoice_number": t.invoice_number,
        "transaction_date": t.transaction_date.isoformat(),
        "notes": t.notes,
        "receipt_url": f"/api/v1/finances/{t.id}/receipt/file" if t.receipt_path else None,
        "created_by": {"id": str(t.created_by.id), "full_name": t.created_by.full_name} if t.created_by else None,
        "created_at": t.created_at.isoformat(),
    }


async def _get_tx(db: AsyncSession, clinic_id: uuid.UUID, tx_id: uuid.UUID) -> FinanceTransaction:
    tx = await db.scalar(
        select(FinanceTransaction).where(
            FinanceTransaction.id == tx_id,
            FinanceTransaction.clinic_id == clinic_id,
        )
    )
    if not tx:
        raise NotFoundError("Transacción")
    return tx


class TransactionCreate(BaseModel):
    type: str
    category: str
    description: str
    original_amount: float
    original_currency: str = "NIO"
    patient_id: str | None = None
    procedure_id: str | None = None
    invoice_number: str | None = None
    transaction_date: date
    notes: str | None = None


class TransactionUpdate(BaseModel):
    category: str | None = None
    description: str | None = None
    original_amount: float | None = None
    original_currency: str | None = None
    patient_id: str | None = None
    procedure_id: str | None = None
    invoice_number: str | None = None
    transaction_date: date | None = None
    notes: str | None = None


class ExchangeRateUpdate(BaseModel):
    rate: float


@router.get("/exchange-rate")
async def get_exchange_rate(
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rate = await finance_service.get_exchange_rate(db, user.clinic_id)
    return {"success": True, "data": {"rate": float(rate)}}


@router.patch("/exchange-rate")
async def update_exchange_rate(
    body: ExchangeRateUpdate,
    user: Annotated[object, require_permission("manage_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rate = await finance_service.set_exchange_rate(db, user.clinic_id, Decimal(str(body.rate)))
    return {"success": True, "data": {"rate": float(rate)}}


@router.get("/summary")
async def get_summary(
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int = Query(...),
    month: int = Query(...),
):
    summary = await finance_service.get_summary(db, user.clinic_id, year, month)
    return {"success": True, "data": summary}


@router.get("/by-patient")
async def income_by_patient(
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int | None = Query(default=None),
    month: int | None = Query(default=None),
):
    data = await finance_service.get_income_by_patient(db, user.clinic_id, year, month)
    return {"success": True, "data": data}


@router.get("/patient/{patient_id_param}")
async def list_patient_transactions(
    patient_id_param: uuid.UUID,
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    txs = await finance_service.list_patient_transactions(db, user.clinic_id, patient_id_param)
    return {"success": True, "data": [_serialize(t) for t in txs]}


@router.get("")
async def list_transactions(
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int = Query(...),
    month: int = Query(...),
    type: str | None = Query(default=None),
):
    txs = await finance_service.list_transactions(db, user.clinic_id, year, month, type)
    return {"success": True, "data": [_serialize(t) for t in txs]}


@router.post("", status_code=201)
async def create_transaction(
    body: TransactionCreate,
    user: Annotated[object, require_permission("manage_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    tx = await finance_service.create_transaction(
        db, user.clinic_id, user.id, body.model_dump()
    )
    return {"success": True, "data": _serialize(tx)}


# ─── Receipt endpoints ─────────────────────────────────────────────────────────

@router.post("/{tx_id}/receipt")
async def upload_receipt(
    tx_id: uuid.UUID,
    user: Annotated[object, require_permission("manage_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
    file: UploadFile = File(...),
):
    if file.content_type not in _ALLOWED_MIME:
        raise ValidationError("Formato no permitido. Use JPEG, PNG, WebP o PDF.")

    data = await file.read()
    if len(data) > _MAX_BYTES:
        raise ValidationError("El archivo excede el tamaño máximo de 10 MB.")

    tx = await _get_tx(db, user.clinic_id, tx_id)

    # Delete old receipt if exists
    if tx.receipt_path:
        storage.delete_file(tx.receipt_path)

    ext = _MIME_TO_EXT[file.content_type]
    path = f"receipts/{user.clinic_id}/{tx_id}.{ext}"
    storage.save_file(path, data)

    tx.receipt_path = path
    await db.commit()
    await db.refresh(tx)

    from app.services.finance_service import _LOAD
    from sqlalchemy.orm import selectinload
    from sqlalchemy import select as sa_select
    result = await db.execute(
        sa_select(FinanceTransaction).where(FinanceTransaction.id == tx_id).options(*_LOAD)
    )
    tx = result.scalar_one()
    return {"success": True, "data": _serialize(tx)}


@router.get("/{tx_id}/receipt/file")
async def get_receipt_file(
    tx_id: uuid.UUID,
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    tx = await _get_tx(db, user.clinic_id, tx_id)
    if not tx.receipt_path:
        raise NotFoundError("Comprobante")
    if not storage.file_exists(tx.receipt_path):
        raise NotFoundError("Archivo de comprobante")
    ext = tx.receipt_path.rsplit(".", 1)[-1].lower()
    mime = {"jpg": "image/jpeg", "png": "image/png", "webp": "image/webp", "pdf": "application/pdf"}.get(ext, "application/octet-stream")
    data = storage.read_file(tx.receipt_path)
    return StreamingResponse(iter([data]), media_type=mime)


@router.delete("/{tx_id}/receipt")
async def delete_receipt(
    tx_id: uuid.UUID,
    user: Annotated[object, require_permission("manage_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    tx = await _get_tx(db, user.clinic_id, tx_id)
    if tx.receipt_path:
        storage.delete_file(tx.receipt_path)
        tx.receipt_path = None
        await db.commit()
    return {"success": True}


@router.patch("/{tx_id}")
async def update_transaction(
    tx_id: uuid.UUID,
    body: TransactionUpdate,
    user: Annotated[object, require_permission("manage_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    tx = await finance_service.update_transaction(
        db, user.clinic_id, tx_id, body.model_dump(exclude_unset=True)
    )
    return {"success": True, "data": _serialize(tx)}


@router.delete("/{tx_id}")
async def delete_transaction(
    tx_id: uuid.UUID,
    user: Annotated[object, require_permission("manage_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    await finance_service.delete_transaction(db, user.clinic_id, tx_id)
    return {"success": True}


@router.get("/export/excel")
async def export_excel(
    user: Annotated[object, require_permission("view_patients")],
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int = Query(...),
    month: int = Query(...),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed"}

    txs = await finance_service.list_transactions(db, user.clinic_id, year, month)
    summary = await finance_service.get_summary(db, user.clinic_id, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Finanzas {year}-{str(month).zfill(2)}"

    MONTHS_ES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill("solid", fgColor="1e3a5f")
    thin = Side(style="thin", color="e2e8f0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Reporte Financiero — {MONTHS_ES[month]} {year}"
    ws["A1"].font = Font(bold=True, size=14, color="1e3a5f")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"] = "RESUMEN DEL MES"
    ws["A2"].font = Font(bold=True, color="FFFFFF")
    ws["A2"].fill = summary_fill
    ws["A2"].alignment = center

    summary_rows = [
        ("Ingresos Brutos", summary["ingresos_brutos"]),
        ("Total Egresos", summary["egresos"]),
        ("Costos Operativos (tratamientos)", summary["costos_operativos"]),
        ("Ingreso Neto (bruto - egresos)", summary["ingreso_neto"]),
        ("Ingreso Neto (incluyendo costos op.)", summary["ingreso_neto_con_op"]),
    ]
    for i, (label, val) in enumerate(summary_rows):
        row = 3 + i
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = val
        ws[f"B{row}"].number_format = '"C$"#,##0.00'
        ws[f"B{row}"].alignment = right_align

    start_row = 9
    headers = ["Fecha", "Tipo", "Categoría", "Descripción", "Paciente",
               "Procedimiento", "Factura #", "Moneda", "Monto Original", "Monto C$", "Costo Op. C$", "Notas"]
    widths = [12, 10, 22, 30, 20, 22, 12, 8, 15, 15, 14, 25]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1e3a5f")
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ALL_CAT = {
        "pago_tratamiento": "Pago tratamiento", "otro_ingreso": "Otro ingreso",
        "laboratorio": "Laboratorio", "insumos": "Insumos", "renta": "Renta",
        "servicios": "Servicios", "salario": "Salario", "otro_egreso": "Otro egreso",
    }

    for i, t in enumerate(txs):
        row = start_row + 1 + i
        row_fill = PatternFill("solid", fgColor="dcfce7") if t.type == "ingreso" else PatternFill("solid", fgColor="fee2e2")
        values = [
            t.transaction_date.strftime("%d/%m/%Y"),
            "Ingreso" if t.type == "ingreso" else "Egreso",
            ALL_CAT.get(t.category, t.category),
            t.description,
            t.patient.full_name if t.patient else "",
            t.procedure.name if t.procedure else "",
            t.invoice_number or "",
            t.original_currency or "NIO",
            float(t.original_amount) if t.original_amount else float(t.amount_cordobas),
            float(t.amount_cordobas),
            float(t.operational_cost_snapshot) if t.operational_cost_snapshot else "",
            t.notes or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.fill = row_fill
            if col in (9, 10, 11):
                cell.number_format = '"C$"#,##0.00'
                cell.alignment = right_align

    total_row = start_row + 1 + len(txs)
    ws.cell(row=total_row, column=9, value="TOTALES").font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=summary["ingresos_brutos"]).font = Font(bold=True, color="16a34a")
    ws.cell(row=total_row, column=10).number_format = '"C$"#,##0.00'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"finanzas_{year}_{str(month).zfill(2)}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
