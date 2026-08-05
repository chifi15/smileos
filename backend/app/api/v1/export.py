import io
import uuid
from datetime import date, datetime
from typing import Annotated

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.dependencies import CurrentUser
from app.models.patient import Patient
from app.models.appointment import Appointment
from app.models.finance import FinanceTransaction
from app.models.costos import CostProduct, CostTreatment, CostProductLot

router = APIRouter(prefix="/export", tags=["Export"])


def _cell(ws, row, col, value, bold=False, fill=None, number_format=None, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF")
    if number_format:
        cell.number_format = number_format
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _header_row(ws, row, cols, fill="1e3a5f"):
    for i, label in enumerate(cols, 1):
        _cell(ws, row, i, label, fill=fill)


def _autofit(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _fmt_date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%d/%m/%Y")
    return str(v)


def _fmt_bool(v) -> str:
    return "Sí" if v else "No"


@router.get("/backup")
async def export_backup(
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    clinic_id = current_user.clinic_id
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── 1. Pacientes ──────────────────────────────────────────────────────────
    result = await db.execute(
        select(Patient)
        .where(Patient.clinic_id == clinic_id)
        .order_by(Patient.patient_number.asc().nulls_last(), Patient.last_name)
    )
    patients = result.scalars().all()

    ws = wb.create_sheet("Pacientes")
    cols = [
        "Nº", "Apellidos", "Nombres", "Cédula", "Fecha nacimiento", "Edad",
        "Género", "Teléfono", "Teléfono 2", "Email", "Ciudad", "País", "Dirección",
        "Ocupación", "Tipo sangre", "Alergias", "Condiciones médicas", "Medicamentos actuales",
        "Motivo de consulta", "Contacto emergencia", "Tel. emergencia",
        "Activo", "Primera visita", "Creado",
    ]
    _header_row(ws, 1, cols)
    for i, p in enumerate(patients, 2):
        age = ""
        if p.date_of_birth:
            today = date.today()
            age = today.year - p.date_of_birth.year - ((today.month, today.day) < (p.date_of_birth.month, p.date_of_birth.day))
        row = [
            p.patient_number, p.last_name, p.first_name, p.id_number,
            _fmt_date(p.date_of_birth), age,
            p.gender, p.phone, p.phone_secondary, p.email,
            p.city, p.country, p.address,
            p.occupation, p.blood_type,
            p.allergies, p.medical_conditions, p.current_medications,
            p.chief_complaint,
            p.emergency_contact_name, p.emergency_contact_phone,
            _fmt_bool(p.is_active), _fmt_date(p.first_visit_date), _fmt_date(p.created_at),
        ]
        for j, val in enumerate(row, 1):
            _cell(ws, i, j, val)
    _autofit(ws)

    # ── 2. Transacciones financieras ──────────────────────────────────────────
    result = await db.execute(
        select(FinanceTransaction)
        .where(FinanceTransaction.clinic_id == clinic_id)
        .options(selectinload(FinanceTransaction.patient), selectinload(FinanceTransaction.procedure))
        .order_by(FinanceTransaction.transaction_date.desc())
    )
    txs = result.scalars().all()

    ws = wb.create_sheet("Finanzas")
    cols = [
        "Fecha", "Tipo", "Categoría", "Descripción", "Paciente",
        "Procedimiento", "Cantidad proc.", "Monto (C$)", "Monto original",
        "Moneda", "Costo operativo", "Nº factura", "Notas", "Creado",
    ]
    _header_row(ws, 1, cols)
    for i, tx in enumerate(txs, 2):
        patient_name = f"{tx.patient.first_name} {tx.patient.last_name}" if tx.patient else ""
        proc_name = tx.procedure.name if tx.procedure else ""
        row = [
            _fmt_date(tx.transaction_date),
            "Ingreso" if tx.type == "ingreso" else "Egreso",
            tx.category, tx.description, patient_name, proc_name,
            tx.procedure_quantity,
            float(tx.amount_cordobas),
            float(tx.original_amount) if tx.original_amount else "",
            tx.original_currency or "",
            float(tx.operational_cost_snapshot) if tx.operational_cost_snapshot else "",
            tx.invoice_number or "", tx.notes or "",
            _fmt_date(tx.created_at),
        ]
        for j, val in enumerate(row, 1):
            c = _cell(ws, i, j, val)
            if j == 8 and val != "":
                c.number_format = '"C$"#,##0.00'
                c.font = Font(color="006400" if tx.type == "ingreso" else "CC0000")
    _autofit(ws)

    # ── 3. Productos / Inventario ─────────────────────────────────────────────
    result = await db.execute(
        select(CostProduct)
        .where(CostProduct.clinic_id == clinic_id)
        .order_by(CostProduct.sort_order)
    )
    products = result.scalars().all()

    ws = wb.create_sheet("Productos")
    cols = [
        "Nombre", "Categoría", "Proveedor", "Presentación", "Unidad",
        "Cant. presentación", "Cant. porción", "Costo total (C$)", "Costo/porción (C$)",
        "Stock actual", "Stock mínimo", "Fecha compra", "Notas",
    ]
    _header_row(ws, 1, cols)
    for i, p in enumerate(products, 2):
        row = [
            p.name, p.category, p.supplier or "",
            p.presentation, p.presentation_unit or "",
            p.presentation_qty, p.portion_qty,
            p.total_cost, p.unit_price,
            p.stock_qty, p.min_stock_qty,
            _fmt_date(p.purchase_date), p.notes or "",
        ]
        for j, val in enumerate(row, 1):
            c = _cell(ws, i, j, val)
            if j in (8, 9) and val:
                c.number_format = '"C$"#,##0.00'
    _autofit(ws)

    # ── 4. Lotes de productos ─────────────────────────────────────────────────
    result = await db.execute(
        select(CostProductLot, CostProduct.name.label("product_name"))
        .join(CostProduct, CostProductLot.product_id == CostProduct.id)
        .where(CostProductLot.clinic_id == clinic_id)
        .order_by(CostProductLot.opened_at.desc())
    )
    lot_rows = result.all()

    ws = wb.create_sheet("Lotes")
    cols = ["Producto", "Abierto", "Cerrado", "Porciones esperadas", "Porciones usadas", "Estado", "Notas"]
    _header_row(ws, 1, cols)
    for i, (lot, product_name) in enumerate(lot_rows, 2):
        estado = "Agotado" if lot.finished_at else "En uso"
        row = [
            product_name,
            _fmt_date(lot.opened_at), _fmt_date(lot.finished_at),
            lot.expected_portions, lot.used_portions,
            estado, lot.notes or "",
        ]
        for j, val in enumerate(row, 1):
            _cell(ws, i, j, val)
    _autofit(ws)

    # ── 5. Tratamientos de costos ─────────────────────────────────────────────
    result = await db.execute(
        select(CostTreatment)
        .where(CostTreatment.clinic_id == clinic_id)
        .options(selectinload(CostTreatment.appointments))
        .order_by(CostTreatment.sort_order)
    )
    treatments = result.scalars().all()

    ws = wb.create_sheet("Tratamientos")
    cols = [
        "Nombre", "Descripción", "Horas totales", "Honorario/hora (C$)",
        "Costos fijos (C$)", "Margen clínica (%)", "Precio sugerido (C$)", "Citas/sesiones",
    ]
    _header_row(ws, 1, cols)
    for i, t in enumerate(treatments, 2):
        row = [
            t.name, t.description or "",
            t.total_hours, t.professional_fee_per_hour,
            t.fixed_costs, t.clinic_margin_pct * 100,
            t.suggested_price, len(t.appointments),
        ]
        for j, val in enumerate(row, 1):
            c = _cell(ws, i, j, val)
            if j in (4, 5, 7) and val:
                c.number_format = '"C$"#,##0.00'
    _autofit(ws)

    # ── 6. Citas ──────────────────────────────────────────────────────────────
    result = await db.execute(
        select(Appointment)
        .where(Appointment.clinic_id == clinic_id)
        .options(selectinload(Appointment.patient))
        .order_by(Appointment.scheduled_at.desc())
    )
    appointments = result.scalars().all()

    ws = wb.create_sheet("Citas")
    cols = ["Fecha", "Hora", "Paciente", "Tipo", "Estado", "Duración (min)", "Razón", "Notas"]
    _header_row(ws, 1, cols)
    for i, apt in enumerate(appointments, 2):
        patient_name = f"{apt.patient.first_name} {apt.patient.last_name}" if apt.patient else ""
        row = [
            _fmt_date(apt.scheduled_at),
            apt.scheduled_at.strftime("%H:%M") if apt.scheduled_at else "",
            patient_name, apt.appointment_type, apt.status,
            apt.duration_minutes, apt.reason or "", apt.notes or "",
        ]
        for j, val in enumerate(row, 1):
            _cell(ws, i, j, val)
    _autofit(ws)

    # ── Portada ───────────────────────────────────────────────────────────────
    ws_cover = wb.create_sheet("Info", 0)
    ws_cover["A1"] = "SmileOS — Respaldo de datos"
    ws_cover["A1"].font = Font(bold=True, size=16, color="1e3a5f")
    ws_cover["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_cover["A2"].font = Font(color="64748b")
    ws_cover["A4"] = "Hojas incluidas:"
    ws_cover["A4"].font = Font(bold=True)
    sheets_info = [
        ("Pacientes", len(patients)),
        ("Finanzas", len(txs)),
        ("Productos", len(products)),
        ("Lotes", len(lot_rows)),
        ("Tratamientos", len(treatments)),
        ("Citas", len(appointments)),
    ]
    for j, (name, count) in enumerate(sheets_info, 5):
        ws_cover[f"A{j}"] = f"  • {name}: {count} registros"
    ws_cover.column_dimensions["A"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"smileos_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/json")
async def export_json(
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    import json
    from app.models.evolution import PatientEvolution

    clinic_id = current_user.clinic_id

    def _iso(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.isoformat()
        if isinstance(v, date):
            return v.isoformat()
        return v

    def _str(v):
        return str(v) if v is not None else None

    # Patients + evolutions
    result = await db.execute(
        select(Patient)
        .where(Patient.clinic_id == clinic_id)
        .order_by(Patient.patient_number.asc().nulls_last())
    )
    patients = result.scalars().all()

    evo_result = await db.execute(
        select(PatientEvolution)
        .where(PatientEvolution.clinic_id == clinic_id)
        .order_by(PatientEvolution.date.desc())
    )
    evolutions = evo_result.scalars().all()
    evos_by_patient: dict[str, list] = {}
    for e in evolutions:
        pid = str(e.patient_id)
        evos_by_patient.setdefault(pid, []).append({
            "id": _str(e.id), "date": _iso(e.date), "note": e.note,
            "attendance": e.attendance, "created_at": _iso(e.created_at),
        })

    patients_data = []
    for p in patients:
        patients_data.append({
            "id": _str(p.id),
            "patient_number": p.patient_number,
            "first_name": p.first_name,
            "last_name": p.last_name,
            "date_of_birth": _iso(p.date_of_birth),
            "gender": p.gender,
            "id_number": p.id_number,
            "occupation": p.occupation,
            "phone": p.phone,
            "phone_secondary": p.phone_secondary,
            "email": p.email,
            "address": p.address,
            "city": p.city,
            "country": p.country,
            "emergency_contact_name": p.emergency_contact_name,
            "emergency_contact_phone": p.emergency_contact_phone,
            "blood_type": p.blood_type,
            "allergies": p.allergies,
            "medical_conditions": p.medical_conditions,
            "current_medications": p.current_medications,
            "chief_complaint": p.chief_complaint,
            "is_active": p.is_active,
            "patient_number": p.patient_number,
            "first_visit_date": _iso(p.first_visit_date),
            "created_at": _iso(p.created_at),
            "evolutions": evos_by_patient.get(str(p.id), []),
        })

    # Transactions
    result = await db.execute(
        select(FinanceTransaction)
        .where(FinanceTransaction.clinic_id == clinic_id)
        .options(selectinload(FinanceTransaction.patient), selectinload(FinanceTransaction.procedure))
        .order_by(FinanceTransaction.transaction_date.desc())
    )
    txs = result.scalars().all()
    transactions_data = []
    for tx in txs:
        transactions_data.append({
            "id": _str(tx.id),
            "type": tx.type,
            "category": tx.category,
            "description": tx.description,
            "amount_cordobas": float(tx.amount_cordobas),
            "original_amount": float(tx.original_amount) if tx.original_amount else None,
            "original_currency": tx.original_currency,
            "exchange_rate_used": float(tx.exchange_rate_used) if tx.exchange_rate_used else None,
            "patient_id": _str(tx.patient_id),
            "patient_name": f"{tx.patient.first_name} {tx.patient.last_name}" if tx.patient else None,
            "procedure_id": _str(tx.procedure_id),
            "procedure_name": tx.procedure.name if tx.procedure else None,
            "procedure_quantity": tx.procedure_quantity,
            "operational_cost_snapshot": float(tx.operational_cost_snapshot) if tx.operational_cost_snapshot else None,
            "invoice_number": tx.invoice_number,
            "transaction_date": _iso(tx.transaction_date),
            "notes": tx.notes,
            "created_at": _iso(tx.created_at),
        })

    # Products + lots
    result = await db.execute(
        select(CostProduct)
        .where(CostProduct.clinic_id == clinic_id)
        .order_by(CostProduct.sort_order)
    )
    products = result.scalars().all()

    lot_result = await db.execute(
        select(CostProductLot)
        .where(CostProductLot.clinic_id == clinic_id)
        .order_by(CostProductLot.opened_at.desc())
    )
    all_lots = lot_result.scalars().all()
    lots_by_product: dict[str, list] = {}
    for lot in all_lots:
        pid = str(lot.product_id)
        lots_by_product.setdefault(pid, []).append({
            "id": _str(lot.id),
            "opened_at": _iso(lot.opened_at),
            "finished_at": _iso(lot.finished_at),
            "expected_portions": lot.expected_portions,
            "used_portions": lot.used_portions,
            "notes": lot.notes,
            "created_at": _iso(lot.created_at),
        })

    products_data = []
    for p in products:
        products_data.append({
            "id": _str(p.id),
            "name": p.name,
            "category": p.category,
            "supplier": p.supplier,
            "notes": p.notes,
            "unit_price": p.unit_price,
            "presentation": p.presentation,
            "portion_description": p.portion_description,
            "total_cost": p.total_cost,
            "presentation_qty": p.presentation_qty,
            "presentation_unit": p.presentation_unit,
            "portion_qty": p.portion_qty,
            "stock_qty": p.stock_qty,
            "min_stock_qty": p.min_stock_qty,
            "purchase_date": _iso(p.purchase_date),
            "sort_order": p.sort_order,
            "created_at": _iso(p.created_at),
            "lots": lots_by_product.get(str(p.id), []),
        })

    # Cost treatments
    result = await db.execute(
        select(CostTreatment)
        .where(CostTreatment.clinic_id == clinic_id)
        .options(selectinload(CostTreatment.appointments))
        .order_by(CostTreatment.sort_order)
    )
    treatments = result.scalars().all()
    treatments_data = []
    for t in treatments:
        treatments_data.append({
            "id": _str(t.id),
            "name": t.name,
            "description": t.description,
            "professional_fee_per_hour": t.professional_fee_per_hour,
            "total_hours": t.total_hours,
            "fixed_costs": t.fixed_costs,
            "clinic_margin_pct": t.clinic_margin_pct,
            "suggested_price": t.suggested_price,
            "procedure_catalog_id": _str(t.procedure_catalog_id),
            "sort_order": t.sort_order,
            "created_at": _iso(t.created_at),
            "appointments": [
                {
                    "id": _str(a.id),
                    "number": a.number,
                    "name": a.name,
                    "materials": a.materials,
                    "sort_order": a.sort_order,
                }
                for a in t.appointments
            ],
        })

    # Appointments
    result = await db.execute(
        select(Appointment)
        .where(Appointment.clinic_id == clinic_id)
        .options(selectinload(Appointment.patient))
        .order_by(Appointment.scheduled_at.desc())
    )
    appointments = result.scalars().all()
    appointments_data = []
    for apt in appointments:
        appointments_data.append({
            "id": _str(apt.id),
            "patient_id": _str(apt.patient_id),
            "patient_name": f"{apt.patient.first_name} {apt.patient.last_name}" if apt.patient else None,
            "scheduled_at": _iso(apt.scheduled_at),
            "duration_minutes": apt.duration_minutes,
            "appointment_type": apt.appointment_type,
            "status": apt.status,
            "reason": apt.reason,
            "notes": apt.notes,
            "created_at": _iso(apt.created_at),
        })

    payload = {
        "export_version": "1.0",
        "generated_at": datetime.now().isoformat(),
        "clinic_id": str(clinic_id),
        "summary": {
            "pacientes": len(patients_data),
            "transacciones": len(transactions_data),
            "productos": len(products_data),
            "tratamientos": len(treatments_data),
            "citas": len(appointments_data),
        },
        "data": {
            "pacientes": patients_data,
            "transacciones": transactions_data,
            "productos": products_data,
            "tratamientos": treatments_data,
            "citas": appointments_data,
        },
    }

    json_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    buf = io.BytesIO(json_bytes)
    filename = f"smileos_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    return StreamingResponse(
        buf,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
